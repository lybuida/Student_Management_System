#export_score_service.py
from StudentManagementApp.models import Teacher, Classroom, Student, AcademicYear
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from StudentManagementApp.dao.score_service import calculate_avg_score


def generate_avg_score_excel(class_id: int, year_id: int, semester: int = None, teacher=None):
    subject_name = teacher.subject.name if teacher else "..."
    classroom = Classroom.query.get(class_id)
    class_name = classroom.name if classroom else "..."
    year_obj = AcademicYear.query.get(year_id)
    academic_year_name = year_obj.name if year_obj else "..."

    students = Student.query.filter_by(classroom_id=class_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"

    setup_headers(ws, subject_name, class_name, academic_year_name, semester)
    populate_score_data(ws, 6, students, year_id, semester)

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"Bang_diem_{subject_name}_{class_name}_{academic_year_name}.xlsx"
    return file_stream, filename


def setup_headers(ws, subject, class_name, year_name, semester):
    merge_range = "A1:C1" if semester else "A1:E1"
    ws.merge_cells(merge_range)
    ws["A1"] = "BẢNG ĐIỂM MÔN HỌC"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Môn học: {subject}"
    ws["A3"] = f"Lớp: {class_name}"
    ws["A4"] = f"Năm học: {year_name}"

    headers = ["STT", "Họ tên"]
    if semester:
        headers.append(f"Điểm TB HK{semester}")
    else:
        headers += ["Điểm TB HK1", "Điểm TB HK2", "Điểm TB cả năm"]

    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(border_style="thin"))

    return header_row


def populate_score_data(ws, header_row, students, academic_year_id, semester):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    left_align = Alignment(horizontal="left")
    center_align = Alignment(horizontal="center")

    for index, student in enumerate(students, 1):
        row = [index, student.name]
        if semester:
            avg = calculate_avg_score(student.id, academic_year_id, semester)
            row.append(round(avg, 2) if avg is not None else "")
        else:
            avg1 = calculate_avg_score(student.id, academic_year_id, 1)
            avg2 = calculate_avg_score(student.id, academic_year_id, 2)
            avg_year = round((avg1 + avg2) / 2, 2) if avg1 is not None and avg2 is not None else ""
            row.extend([
                round(avg1, 2) if avg1 is not None else "",
                round(avg2, 2) if avg2 is not None else "",
                avg_year
            ])
        ws.append(row)

    for i, column_cells in enumerate(ws.iter_cols(min_row=header_row, max_row=ws.max_row), 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.border = thin_border
            cell.alignment = center_align if i != 2 else left_align
            cell.protection = cell.protection.copy(locked=True)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.protection.sheet = True
    ws.protection.password = '123'
